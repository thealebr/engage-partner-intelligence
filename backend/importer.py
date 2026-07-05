import argparse, hashlib, re, sqlite3, unicodedata
from pathlib import Path
from openpyxl import load_workbook

SPECIALIZATION_COLUMNS = {
    'SD-WAN Eligible?': 'sd_wan', 'SASE Eligible?': 'sase',
    'Secure Networking LAN Eligible?': 'secure_networking_lan',
    'Cloud Security Eligible?': 'cloud_security',
    'Secure Networking Firewall Eligible?': 'secure_networking_firewall',
    'Security Operations Eligible?': 'security_operations', 'OT Eligible?': 'ot'
}
def norm(value):
    text=unicodedata.normalize('NFD',str(value or '')).encode('ascii','ignore').decode().lower()
    return re.sub(r'[^a-z0-9]+',' ',text).strip()
def truth(value): return str(value).strip().lower() in {'true','1','yes','sim','y'}
def sha(path): return hashlib.sha256(Path(path).read_bytes()).hexdigest()
def validate_metadata(path, owner, year, quarter, dataset_type):
    wb=load_workbook(path,data_only=True,read_only=True)
    sheet=next((s for s in wb.worksheets if s.title.strip().upper()=='IMPORT_META'),None)
    if sheet is None: raise ValueError(f'{Path(path).name}: aba IMPORT_META não encontrada')
    values={str(row[0] or '').strip().upper().replace(' ','_'):str(row[1] or '').strip() for row in sheet.iter_rows(min_row=1,max_col=2,values_only=True) if row[0]}
    required={'ACCOUNT_OWNER','YEAR','QUARTER','DATASET_TYPE'}
    missing=required-values.keys()
    if missing: raise ValueError(f'{Path(path).name}: campos ausentes na IMPORT_META: {", ".join(sorted(missing))}')
    actual_quarter=values['QUARTER'].upper().replace('Q','')
    if values['ACCOUNT_OWNER']!=owner: raise ValueError(f'{Path(path).name}: Account Owner da planilha não corresponde ao formulário')
    if int(float(values['YEAR']))!=year: raise ValueError(f'{Path(path).name}: Ano da planilha não corresponde ao formulário')
    if int(float(actual_quarter))!=quarter: raise ValueError(f'{Path(path).name}: Quarter da planilha não corresponde ao formulário')
    if values['DATASET_TYPE'].upper()!=dataset_type: raise ValueError(f'{Path(path).name}: DATASET_TYPE deve ser {dataset_type}')
def rows(path):
    wb=load_workbook(path,data_only=True,read_only=True)
    ws=next((s for s in wb.worksheets if s.title.strip().upper()!='IMPORT_META' and 'Account Name' in [str(v).strip() if v is not None else '' for v in next(s.iter_rows(max_row=1,values_only=True))]),None)
    if ws is None: raise ValueError(f'{Path(path).name}: aba de dados não encontrada')
    headers=[str(v).strip() if v is not None else '' for v in next(ws.iter_rows(values_only=True))]
    for values in ws.iter_rows(values_only=True):
        record=dict(zip(headers,values))
        if any(v is not None for v in values): yield record

def dedupe_engage(path, owner):
    output={}
    for r in rows(path):
        if str(r.get('Account Owner') or '').strip()!=owner or not r.get('Account Name'): continue
        item={'name':str(r['Account Name']).strip(),'level':str(r.get('Level of Engagement') or '').strip(),
              'integrator':truth(r.get('Fully Compliant (Integrator)')),'mssp_account':truth(r.get('MSSP Account')),
              'mssp':truth(r.get('Fully Compliant (MSSP)'))}
        key=norm(item['name'])
        if key in output and output[key]!=item: raise ValueError(f'Valores conflitantes para {item["name"]} na planilha Engage')
        output[key]=item
    return output

def dedupe_specialize(path, owner):
    output={}
    for r in rows(path):
        if str(r.get('Account Owner') or '').strip()!=owner or not r.get('Account Name'): continue
        item={'name':str(r['Account Name']).strip(), **{db:truth(r.get(xls)) for xls,db in SPECIALIZATION_COLUMNS.items()}}
        key=norm(item['name'])
        if key in output and output[key]!=item: raise ValueError(f'Valores conflitantes para {item["name"]} na planilha Specialize')
        output[key]=item
    return output

def import_pair(db, engage, specialize, owner, year, quarter):
    validate_metadata(engage,owner,year,quarter,'ENGAGE_COMPLIANCE')
    validate_metadata(specialize,owner,year,quarter,'SPECIALIZATIONS')
    engage_rows=dedupe_engage(engage,owner); spec_rows=dedupe_specialize(specialize,owner)
    if not engage_rows: raise ValueError('Nenhum canal encontrado na planilha Engage para este Account Owner')
    if not spec_rows: raise ValueError('Nenhum canal encontrado na planilha Specialize para este Account Owner')
    con=sqlite3.connect(db); con.execute('PRAGMA foreign_keys=ON')
    con.executescript(Path(__file__).with_name('schema.sql').read_text(encoding='utf-8-sig'))
    try:
        con.execute('BEGIN')
        version=con.execute('SELECT COALESCE(MAX(version),0)+1 FROM import_batches WHERE account_owner=? AND year=? AND quarter=?',(owner,year,quarter)).fetchone()[0]
        con.execute("UPDATE import_batches SET is_active=0,status='superseded' WHERE account_owner=? AND year=? AND quarter=? AND is_active=1",(owner,year,quarter))
        cur=con.execute('INSERT INTO import_batches(account_owner,year,quarter,version,engage_file,specialize_file,engage_sha256,specialize_sha256,status,is_active) VALUES(?,?,?,?,?,?,?,?,?,1)',(owner,year,quarter,version,Path(engage).name,Path(specialize).name,sha(engage),sha(specialize),'complete'))
        batch=cur.lastrowid
        partner_ids={}
        for key,item in {**spec_rows,**engage_rows}.items():
            con.execute('INSERT INTO partners(canonical_name,normalized_name) VALUES(?,?) ON CONFLICT(normalized_name) DO UPDATE SET canonical_name=excluded.canonical_name',(item['name'],key))
            partner_ids[key]=con.execute('SELECT id FROM partners WHERE normalized_name=?',(key,)).fetchone()[0]
        for key,item in engage_rows.items():
            con.execute('INSERT INTO partner_snapshots VALUES(?,?,?,?,?,?)',(batch,partner_ids[key],item['level'],item['integrator'],item['mssp_account'],item['mssp']))
        fields=list(SPECIALIZATION_COLUMNS.values())
        for key,item in spec_rows.items():
            con.execute('INSERT INTO specialization_snapshots VALUES('+','.join('?' for _ in range(9))+')',(batch,partner_ids[key],*[item[f] for f in fields]))
        con.commit(); return {'batch_id':batch,'version':version,'engage_partners':len(engage_rows),'specialization_partners':len(spec_rows)}
    except: con.rollback(); raise
    finally: con.close()

if __name__=='__main__':
    p=argparse.ArgumentParser(); p.add_argument('--db',default='engage.db'); p.add_argument('--engage',required=True); p.add_argument('--specialize',required=True); p.add_argument('--owner',required=True); p.add_argument('--year',type=int,required=True); p.add_argument('--quarter',type=int,choices=range(1,5),required=True); a=p.parse_args()
    print(import_pair(a.db,a.engage,a.specialize,a.owner,a.year,a.quarter))


EXAM_COLUMNS={'FCF - Cybersecurity #':'fcf','FCA - Cybersecurity #':'fca','FCP - Any #':'fcp','FCSS - Any #':'fcss'}
def dedupe_exams(path,owner):
    output={}
    for r in rows(path):
        if str(r.get('Account Owner') or '').strip()!=owner or not r.get('Account Name'): continue
        item={'name':str(r['Account Name']).strip(),'level':str(r.get('Level of Engagement') or '').strip(),**{field:int(r.get(column) or 0) for column,field in EXAM_COLUMNS.items()}}
        key=norm(item['name'])
        if key in output and output[key]!=item: raise ValueError(f'Valores de exames conflitantes para {item["name"]}')
        output[key]=item
    return output

def import_exams(db,exam_file,owner,year,quarter,validate_meta=True):
    if validate_meta: validate_metadata(exam_file,owner,year,quarter,'COMPLIANCE_EXAMS')
    exam_rows=dedupe_exams(exam_file,owner)
    if not exam_rows: raise ValueError('Nenhum canal encontrado na planilha de exames para este Account Owner')
    con=sqlite3.connect(db); con.execute('PRAGMA foreign_keys=ON'); con.executescript(Path(__file__).with_name('schema.sql').read_text(encoding='utf-8-sig'))
    try:
        con.execute('BEGIN'); version=con.execute('SELECT COALESCE(MAX(version),0)+1 FROM exam_import_batches WHERE account_owner=? AND year=? AND quarter=?',(owner,year,quarter)).fetchone()[0]
        con.execute("UPDATE exam_import_batches SET is_active=0,status='superseded' WHERE account_owner=? AND year=? AND quarter=? AND is_active=1",(owner,year,quarter))
        cur=con.execute('INSERT INTO exam_import_batches(account_owner,year,quarter,version,source_file,source_sha256,status,is_active) VALUES(?,?,?,?,?,?,?,1)',(owner,year,quarter,version,Path(exam_file).name,sha(exam_file),'complete')); batch=cur.lastrowid
        imported=0; missing=[]
        for key,item in exam_rows.items():
            partner=con.execute('SELECT id FROM partners WHERE normalized_name=?',(key,)).fetchone()
            if not partner: missing.append(item['name']); continue
            con.execute('INSERT INTO exam_snapshots VALUES(?,?,?,?,?,?,?)',(batch,partner[0],item['level'],item['fcf'],item['fca'],item['fcp'],item['fcss'])); imported+=1
        if missing: raise ValueError('Canais não encontrados no banco: '+', '.join(missing[:10]))
        con.commit(); return {'batch_id':batch,'version':version,'exam_partners':imported}
    except: con.rollback(); raise
    finally: con.close()
